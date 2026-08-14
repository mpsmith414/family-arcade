#!/usr/bin/env python3
"""Build the Kling reference-prompt superhero character sheet workbook."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.environ.get("OUT", "kling-superhero-character-sheet.xlsx")

# ---------------------------------------------------------------- palette
INK        = "1F2430"
NAVY       = "1F3864"
BAND       = "2E5395"
SECTION    = "D9E2F3"
INPUT_FILL = "FFF2CC"   # cells the user fills in
LOCK_FILL  = "F2F2F2"
WHITE      = "FFFFFF"

F  = "Arial"
def font(sz=10, b=False, i=False, color=INK):
    return Font(name=F, size=sz, bold=b, italic=i, color=color)

TITLE_F   = font(16, True, color=WHITE)
BAND_F    = font(11, True, color=WHITE)
SECT_F    = font(10, True, color=NAVY)
BODY_F    = font(10)
INPUT_F   = font(10, color="0000FF")      # blue = hardcoded input
FORMULA_F = font(10)                       # black = formula
NOTE_F    = font(9, i=True, color="595959")

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL  = PatternFill("solid", fgColor=BAND)
SECT_FILL  = PatternFill("solid", fgColor=SECTION)
IN_FILL    = PatternFill("solid", fgColor=INPUT_FILL)
GREY_FILL  = PatternFill("solid", fgColor=LOCK_FILL)

thin = Side(style="thin", color="BFBFBF")
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

TOP  = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
MID  = Alignment(horizontal="left", vertical="center", wrap_text=True)


def title_row(ws, text, span, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.fill, c.alignment = TITLE_F, TITLE_FILL, MID
    ws.row_dimensions[row].height = 30


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# =====================================================================
# 1. READ ME
# =====================================================================
ws = wb.active
ws.title = "Read Me"
widths(ws, {"A": 4, "B": 26, "C": 96})
title_row(ws, "Kling Reference Prompts — Superhero Character Sheet", 3)

r = 3
ws.cell(row=r, column=2, value="What this is").font = SECT_F
ws.cell(row=r, column=3, value=(
    "A prompt kit for Kling AI. The Character Sheet tab holds one column per hero — every "
    "detail that has to stay identical from shot to shot. The Prompt Builder assembles those "
    "details into finished prompt strings, so a costume change is made once and every prompt "
    "updates. The Shot List turns a finished reference image into a sequence of video shots."
)).alignment = TOP
ws.row_dimensions[r].height = 58

r += 2
ws.cell(row=r, column=2, value="Workflow").font = SECT_F
steps = [
    ("1. Fill the Character Sheet",
     "One column per hero. Every yellow cell. Be concrete — 'cobalt blue' beats 'blue', "
     "'shoulder-length curly' beats 'curly'. Vague fields are where Kling improvises, and "
     "improvisation is what breaks continuity between shots."),
    ("2. Pick a hero",
     "Prompt Builder tab, cell B2 — a drop-down of the names in row 3 of the Character Sheet. "
     "Every prompt below it re-resolves to that hero."),
    ("3. Generate the reference image",
     "Run the Master Reference prompt in Kling's image generator first. Regenerate until it is "
     "right. This single image is the anchor for everything after it — do not skip ahead."),
    ("4. Lock continuity",
     "Save that image, then record its file name and seed / image ID back in the Character "
     "Sheet's Continuity block. That is what lets you come back in a month and match."),
    ("5. Animate",
     "Feed the reference image into Kling image-to-video and use the Shot List prompts for "
     "motion. Keep clips to 5s unless a shot genuinely needs 10s — longer clips drift more."),
]
for label, body in steps:
    r += 1
    c = ws.cell(row=r, column=2, value=label); c.font = font(10, True); c.alignment = TOP
    c = ws.cell(row=r, column=3, value=body);  c.font = BODY_F; c.alignment = TOP
    ws.row_dimensions[r].height = 44

r += 2
ws.cell(row=r, column=2, value="Colour legend").font = SECT_F
legend = [
    ("Yellow fill, blue text", "You type here. These are the only cells meant to be edited."),
    ("Black text, no fill",    "A formula. It builds itself from the yellow cells — don't overwrite it."),
    ("Grey fill",              "Reference library. Copy from it; nothing here feeds a formula."),
]
for label, body in legend:
    r += 1
    c = ws.cell(row=r, column=2, value=label); c.font = font(10, True); c.alignment = TOP
    if label.startswith("Yellow"):
        c.fill = IN_FILL
    elif label.startswith("Grey"):
        c.fill = GREY_FILL
    c = ws.cell(row=r, column=3, value=body); c.font = BODY_F; c.alignment = TOP
    ws.row_dimensions[r].height = 18

r += 2
ws.cell(row=r, column=2, value="Prompt anatomy").font = SECT_F
ws.cell(row=r, column=3, value=(
    "Kling reads a prompt front-to-back and weights the front more heavily. The builder writes "
    "every prompt in the same order for that reason:\n"
    "    SUBJECT  →  APPEARANCE  →  COSTUME  →  ACTION  →  CAMERA  →  LIGHTING  →  STYLE\n"
    "A video prompt should describe motion and nothing else the reference image already shows. "
    "Re-describing the costume in a video prompt invites Kling to redesign it."
)).alignment = TOP
ws.row_dimensions[r].height = 62

r += 2
ws.cell(row=r, column=2, value="Assumptions").font = SECT_F
ws.cell(row=r, column=3, value=(
    "The two worked examples are named for the two characters in this repo's games, as a "
    "format demo — overwrite row 3 of the Character Sheet to rename them. Kling settings named "
    "on the Shot List (5s/10s durations, 16:9 / 9:16 / 1:1 aspect ratios, Standard vs "
    "Professional mode) reflect Kling's published options; check them against the version you "
    "are on, since the model line moves quickly."
)).alignment = TOP
ws.row_dimensions[r].height = 62

ws.sheet_view.showGridLines = False

# =====================================================================
# 2. CHARACTER SHEET
# =====================================================================
cs = wb.create_sheet("Character Sheet")
widths(cs, {"A": 26, "B": 40, "C": 40, "D": 40, "E": 46})
title_row(cs, "Character Sheet — the details that must never drift", 5)

cs.cell(row=2, column=1, value=(
    "One column per hero. Fill every yellow cell; a blank field is a field Kling invents "
    "differently in every shot."
)).font = NOTE_F
cs.row_dimensions[2].height = 16

HDR = 3
for col, val in enumerate(["Field", "Comet Kid", "Sparkfish", "Hero 3", "Why it matters"], start=1):
    c = cs.cell(row=HDR, column=col, value=val)
    c.font, c.fill, c.alignment, c.border = BAND_F, BAND_FILL, CTR, BOX
cs.row_dimensions[HDR].height = 22

# (section, field, char1, char2, note)
FIELDS = [
    ("IDENTITY", None, None, None),
    ("Played by / based on", "Oliver", "Emsile",
     "Name the real person or leave blank. Kling never sees this field — it is your note."),
    ("Age / apparent age", "7 years old", "5 years old",
     "Kling defaults to adult heroes. Say the age or you get a grown-up in a small suit."),
    ("One-line concept", "a fearless boy speedster who leaves a comet trail",
     "a bright girl who talks to sea creatures and rides a jet of water",
     "The elevator pitch. Front-loads the prompt and steers everything after it."),
    ("Build / height", "small and wiry, child proportions", "petite, rounded child proportions",
     "'child proportions' is the single most load-bearing phrase in the whole sheet."),

    ("FACE", None, None, None),
    ("Face shape", "round face, freckles across the nose", "round face, dimpled cheeks",
     "Two concrete markers are enough. Ten makes the face mushy."),
    ("Skin tone", "fair skin", "fair skin",
     "State it. Left out, it changes shot to shot."),
    ("Hair colour", "sandy blond", "light brown",
     "Name a colour a paint chart would recognise."),
    ("Hair style", "short and messy, swept forward", "shoulder-length, two braids",
     "Length and shape. Braids and ponytails survive motion better than loose hair."),
    ("Eye colour", "bright blue", "hazel", "One word. Reinforced in every prompt."),
    ("Expression at rest", "grinning, eager", "calm, wide-eyed curiosity",
     "The neutral face. Anything else you ask for is a departure from this."),

    ("COSTUME", None, None, None),
    ("Suit base colour", "deep cobalt blue", "seafoam green",
     "Be specific. 'Blue' spans navy to sky and Kling will roam the whole range."),
    ("Secondary colour", "white side panels", "coral pink side panels",
     "The second colour blocks out the silhouette."),
    ("Accent / trim colour", "gold piping", "silver piping",
     "Trim reads at a distance and is the first thing to go missing. Keep it simple."),
    ("Material / finish", "matte stretch fabric", "smooth satin finish",
     "Matte survives lighting changes; gloss throws highlights that move between shots."),
    ("Chest emblem", "a gold comet streaking across a white circle",
     "a silver fish leaping over a coral wave",
     "Keep it to one shape and two colours. Detailed emblems re-draw themselves every time."),
    ("Cape / cloak", "short gold cape to the waist", "no cape",
     "Write 'no cape' rather than leaving it blank — blank is an invitation."),
    ("Mask / headgear", "blue domino mask", "coral headband, no mask",
     "A mask that covers the eyes will fight any expression you ask for."),
    ("Gloves", "white gloves to the forearm", "short coral gloves", "Length matters at the wrist."),
    ("Boots", "white boots with gold soles", "coral boots with silver soles",
     "Boots anchor the bottom of the silhouette."),
    ("Belt / utility", "gold belt with a round buckle", "silver belt with a shell buckle",
     "One belt detail. Pouches and gadgets multiply on their own."),

    ("POWERS & FX", None, None, None),
    ("Primary power", "super speed", "water control",
     "One power. Two powers halve the clarity of both."),
    ("VFX colour", "gold and white", "aqua and white",
     "Tie the effect to the trim colour and the character reads as one design."),
    ("VFX shape", "a trailing comet streak of sparks", "arcing ribbons of water",
     "Shape beats intensity. 'Streak' and 'ribbon' animate; 'glow' just sits there."),
    ("Signature prop", "none", "a small glowing conch shell",
     "Optional, but a prop gives the hands something to do in a video shot."),
    ("Power pose", "crouched low, one fist forward, ready to sprint",
     "arms raised, water spiralling upward around her",
     "The pose for the hero shot. Should be readable in silhouette."),

    ("PERFORMANCE", None, None, None),
    ("Personality", "brave, impatient, funny", "gentle, watchful, brave when it counts",
     "Three adjectives. Feeds the face and the body language, not the costume."),
    ("Movement style", "quick, bouncy, always leaning forward",
     "smooth and flowing, light on her feet",
     "This is the field that makes video shots feel like the character."),

    ("CONTINUITY LOCK", None, None, None),
    ("Art style", "bright 3D animated film style, Pixar-like, soft rounded shapes",
     "bright 3D animated film style, Pixar-like, soft rounded shapes",
     "Identical across every hero in one project, or they look like guests in each other's films."),
    ("Reference image file", "comet-kid-ref-01.png", "sparkfish-ref-01.png",
     "Fill in AFTER you generate the master reference. This is the anchor image."),
    ("Seed / image ID", "(paste after generating)", "(paste after generating)",
     "Kling's seed or image ID. Without it a matching re-generation is luck."),
    ("Aspect ratio", "16:9", "16:9", "Keep one ratio per project. Switching re-frames the character."),
    ("Never change", "emblem shape, cape length, mask colour",
     "emblem shape, braid count, headband colour",
     "Your own checklist. Read it against each new render before you accept it."),
]

row = HDR + 1
FIELD_ROW = {}
for entry in FIELDS:
    name, v1, v2, note = entry
    if v1 is None:  # section band
        cs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = cs.cell(row=row, column=1, value=name)
        c.font, c.fill, c.alignment = SECT_F, SECT_FILL, MID
        cs.row_dimensions[row].height = 18
        row += 1
        continue

    c = cs.cell(row=row, column=1, value=name)
    c.font, c.alignment, c.border = font(10, True), TOP, BOX
    for col, val in ((2, v1), (3, v2)):
        c = cs.cell(row=row, column=col, value=val)
        c.font, c.fill, c.alignment, c.border = INPUT_F, IN_FILL, TOP, BOX
    c = cs.cell(row=row, column=4)                       # blank third hero
    c.fill, c.alignment, c.border, c.font = IN_FILL, TOP, BOX, INPUT_F
    c = cs.cell(row=row, column=5, value=note)
    c.font, c.alignment, c.border = NOTE_F, TOP, BOX
    cs.row_dimensions[row].height = 30
    FIELD_ROW[name] = row
    row += 1

FIRST_FIELD, LAST_FIELD = HDR + 1, row - 1
cs.freeze_panes = "B4"
cs.sheet_view.showGridLines = False

# =====================================================================
# 3. PROMPT BUILDER
# =====================================================================
pb = wb.create_sheet("Prompt Builder")
widths(pb, {"A": 26, "B": 104, "C": 34})
title_row(pb, "Prompt Builder — pick a hero, copy a prompt", 3)

pb.cell(row=2, column=1, value="Hero:").font = font(11, True)
sel = pb.cell(row=2, column=2, value="Comet Kid")
sel.font, sel.fill, sel.border, sel.alignment = font(11, True, color="0000FF"), IN_FILL, BOX, MID
pb.cell(row=2, column=3, value="← drop-down; the whole sheet re-resolves").font = NOTE_F
pb.row_dimensions[2].height = 20

dv = DataValidation(type="list", formula1="'Character Sheet'!$B$3:$D$3", allow_blank=False)
pb.add_data_validation(dv)
dv.add(sel)

# --- resolved attributes -------------------------------------------------
r = 4
pb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = pb.cell(row=r, column=1, value="RESOLVED ATTRIBUTES  (INDEX/MATCH into the Character Sheet)")
c.font, c.fill, c.alignment = BAND_F, BAND_FILL, MID
r += 1

RES = {}
for name in FIELD_ROW:
    c = pb.cell(row=r, column=1, value=name)
    c.font, c.alignment, c.border = font(10, True), MID, BOX
    f = (f"=INDEX('Character Sheet'!$B${FIRST_FIELD}:$D${LAST_FIELD},"
         f"MATCH($A{r},'Character Sheet'!$A${FIRST_FIELD}:$A${LAST_FIELD},0),"
         f"MATCH($B$2,'Character Sheet'!$B$3:$D$3,0))")
    c = pb.cell(row=r, column=2, value=f)
    c.font, c.alignment, c.border = FORMULA_F, MID, BOX
    RES[name] = f"$B${r}"
    r += 1

R = RES  # shorthand


def q(s):
    """A literal string fragment inside a concatenation formula."""
    return '"' + s.replace('"', '""') + '"'


def build(parts):
    """parts: list of literal strings and $B$n cell refs -> one & formula."""
    out = []
    for p in parts:
        out.append(p if p.startswith("$") else q(p))
    return "=" + "&".join(out)


ALIAS = "$B$2"

# --- built prompts -------------------------------------------------------
r += 1
pb.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
c = pb.cell(row=r, column=1, value="BUILT PROMPTS  (copy column B straight into Kling)")
c.font, c.fill, c.alignment = BAND_F, BAND_FILL, MID
r += 1

for col, val in enumerate(["Prompt", "Text to copy", "Use it for"], start=1):
    c = pb.cell(row=r, column=col, value=val)
    c.font, c.fill, c.alignment, c.border = font(10, True, color=NAVY), SECT_FILL, CTR, BOX
r += 1

look = [
    ", ", R["Build / height"], ", ", R["Face shape"], ", ", R["Skin tone"], ", ",
    R["Hair style"], " ", R["Hair colour"], " hair, ", R["Eye colour"], " eyes",
]
suit = [
    ", wearing a ", R["Material / finish"], " ", R["Suit base colour"], " superhero suit with ",
    R["Secondary colour"], " and ", R["Accent / trim colour"], ", ", R["Chest emblem"],
    " on the chest, ", R["Cape / cloak"], ", ", R["Mask / headgear"], ", ", R["Gloves"],
    ", ", R["Boots"], ", ", R["Belt / utility"],
]

PROMPTS = [
    ("1 · Master reference",
     build([ALIAS, ", ", R["One-line concept"], *look, *suit,
            ". Full body, standing straight in a relaxed A-pose, facing camera, arms slightly "
            "away from the body, both hands visible, ", R["Expression at rest"],
            ". Plain light grey studio background, soft even three-point lighting, no shadows "
            "on the backdrop, sharp focus, entire figure inside the frame with headroom. ",
            R["Art style"], "."]),
     "Generate this FIRST. Everything else keys off the image it produces."),

    ("2 · Turnaround sheet",
     build(["character model sheet of ", ALIAS, *look, *suit,
            ". Four full-body views in a row on one image: front, three-quarter, side, back. "
            "Identical pose, identical scale and identical costume in all four. Plain white "
            "background, flat even lighting, no shadows, orthographic framing. ",
            R["Art style"], "."]),
     "The continuity bible. Print it and check every later render against it."),

    ("3 · Expression sheet",
     build(["expression sheet of ", ALIAS, ", head and shoulders only, ", R["Face shape"],
            ", ", R["Hair style"], " ", R["Hair colour"], " hair, ", R["Eye colour"], " eyes, ",
            R["Mask / headgear"],
            ". Six portraits in a two by three grid: happy, surprised, determined, laughing, "
            "worried, sleepy. Same face, same lighting and same angle in every panel. Plain "
            "white background. ", R["Art style"], "."]),
     "Do this before any dialogue or reaction shot."),

    ("4 · Costume detail",
     build(["costume detail sheet for ", ALIAS, ": close-up studies of ", R["Chest emblem"],
            ", ", R["Gloves"], ", ", R["Boots"], ", and ", R["Belt / utility"],
            ", all in ", R["Material / finish"], " ", R["Suit base colour"], " with ",
            R["Accent / trim colour"],
            ". Laid out on a plain white background with even lighting, no character present. ",
            R["Art style"], "."]),
     "Settles the small parts that otherwise redraw themselves each shot."),

    ("5 · Hero pose",
     build([ALIAS, ", ", R["One-line concept"], *look, *suit, ". ", R["Power pose"],
            ", ", R["VFX shape"], " in ", R["VFX colour"], " around ", ALIAS,
            ". Low angle hero shot, wide lens, dramatic rim light from behind, "
            "city rooftop at golden hour, shallow depth of field. ", R["Art style"], "."]),
     "The poster shot. Also the best still to animate first."),

    ("6 · Video base (image-to-video)",
     build([ALIAS, " moves in character: ", R["Movement style"], ", with ", R["VFX shape"], " in ", R["VFX colour"],
            " streaming behind. Camera slowly pushes in. The character, the costume and the "
            "colours stay exactly as in the reference image. Smooth natural motion, "
            "consistent lighting throughout."]),
     "Paste the reference image in, then this text. Describes motion only — never the costume."),

    ("7 · Two-hero shot",
     build([ALIAS, " and a second hero standing side by side, both in full costume, matched "
            "scale and matched lighting, facing camera, ", R["Expression at rest"],
            ". Plain light grey studio background, soft even lighting, both figures fully in "
            "frame. ", R["Art style"], "."]),
     "Feed BOTH reference images. Match scale here or one hero towers over the other."),
]

for name, formula, use in PROMPTS:
    c = pb.cell(row=r, column=1, value=name)
    c.font, c.alignment, c.border = font(10, True), TOP, BOX
    c = pb.cell(row=r, column=2, value=formula)
    c.font, c.alignment, c.border = FORMULA_F, TOP, BOX
    c = pb.cell(row=r, column=3, value=use)
    c.font, c.alignment, c.border = NOTE_F, TOP, BOX
    pb.row_dimensions[r].height = 96
    r += 1

r += 1
c = pb.cell(row=r, column=1, value="Negative prompt")
c.font, c.alignment, c.border = font(10, True), TOP, BOX
c = pb.cell(row=r, column=2, value="='Negative Prompts'!$B$4")
c.font, c.alignment, c.border = FORMULA_F, TOP, BOX
c = pb.cell(row=r, column=3, value="Paste into Kling's negative field alongside any prompt above.")
c.font, c.alignment, c.border = NOTE_F, TOP, BOX
pb.row_dimensions[r].height = 46

pb.freeze_panes = "A5"
pb.sheet_view.showGridLines = False

# =====================================================================
# 4. SHOT LIST
# =====================================================================
sl = wb.create_sheet("Shot List")
widths(sl, {"A": 6, "B": 22, "C": 24, "D": 26, "E": 78, "F": 9, "G": 9, "H": 14, "I": 12})
title_row(sl, "Shot List — image-to-video, one row per clip", 9)
sl.cell(row=2, column=1, value=(
    "Column E builds itself from the hero picked on Prompt Builder plus the yellow cells in C "
    "and D. Motion only: say nothing here that the reference image already shows."
)).font = NOTE_F
sl.row_dimensions[2].height = 16

SHDR = 3
for col, val in enumerate(
        ["#", "Beat", "Action", "Camera", "Prompt (built)", "Secs", "Ratio", "Mode", "Status"],
        start=1):
    c = sl.cell(row=SHDR, column=col, value=val)
    c.font, c.fill, c.alignment, c.border = BAND_F, BAND_FILL, CTR, BOX
sl.row_dimensions[SHDR].height = 22

SHOTS = [
    ("Establish", "stands still and looks up at the sky, cape and hair moving in the breeze",
     "slow push in, eye level", 5, "16:9", "Standard"),
    ("Turn to camera", "turns their head to camera and breaks into a grin",
     "static close-up, eye level", 5, "16:9", "Professional"),
    ("Power up", "raises both arms as the effect builds and swirls around them",
     "slow orbit to the left, low angle", 5, "16:9", "Professional"),
    ("Take off", "crouches, then launches upward out of the top of the frame",
     "static wide, low angle, tilts up to follow", 5, "16:9", "Professional"),
    ("Flight", "flies toward camera, effect streaming behind, clouds rushing past",
     "camera tracks backwards ahead of them", 10, "16:9", "Professional"),
    ("Landing", "lands hard in a crouch, one fist to the ground, dust ring spreading outward",
     "static low angle wide, slight shake on impact", 5, "16:9", "Professional"),
    ("Rescue", "catches a falling object, hugs it safe to their chest and looks relieved",
     "medium shot, slow push in", 5, "16:9", "Professional"),
    ("Team up", "the two heroes bump fists and turn to face camera together",
     "static medium two-shot, eye level", 5, "16:9", "Professional"),
    ("Victory", "laughs and punches the air, effect bursting outward",
     "slow pull back to a wide shot", 5, "16:9", "Standard"),
    ("Walk away", "walks away from camera into the light, cape swinging",
     "static wide, low angle", 10, "16:9", "Professional"),
    ("Phone cut", "poses, arms folded, effect glowing steadily around them",
     "slow push in, eye level", 5, "9:16", "Standard"),
    ("Title card", "stands motionless as the effect drifts slowly across the frame",
     "very slow drift to the right", 5, "1:1", "Standard"),
]

r = SHDR + 1
for i, (beat, action, camera, secs, ratio, mode) in enumerate(SHOTS, start=1):
    c = sl.cell(row=r, column=1, value=i)
    c.font, c.alignment, c.border = BODY_F, CTR, BOX
    c = sl.cell(row=r, column=2, value=beat)
    c.font, c.alignment, c.border = font(10, True), TOP, BOX
    for col, val in ((3, action), (4, camera)):
        c = sl.cell(row=r, column=col, value=val)
        c.font, c.fill, c.alignment, c.border = INPUT_F, IN_FILL, TOP, BOX
    f = ("=\"The character from the reference image \"&$C{r}&\", \"&$D{r}"
         "&\". \"&'Prompt Builder'!$B$2&\" moves in character: \"&'Prompt Builder'!{mv}"
         "&\". Costume, colours and face stay exactly as the reference image. \""
         "&\"Smooth natural motion, consistent lighting, no cuts.\"").format(
            r=r, mv=RES["Movement style"])
    c = sl.cell(row=r, column=5, value=f)
    c.font, c.alignment, c.border = FORMULA_F, TOP, BOX
    for col, val in ((6, secs), (7, ratio), (8, mode)):
        c = sl.cell(row=r, column=col, value=val)
        c.font, c.alignment, c.border = BODY_F, CTR, BOX
    c = sl.cell(row=r, column=9, value="To do")
    c.font, c.fill, c.alignment, c.border = INPUT_F, IN_FILL, CTR, BOX
    sl.row_dimensions[r].height = 58
    r += 1

TOT = r
c = sl.cell(row=TOT, column=5, value="Total runtime (seconds)")
c.font, c.alignment, c.border = font(10, True), Alignment(horizontal="right", vertical="center"), BOX
c = sl.cell(row=TOT, column=6, value=f"=SUM(F{SHDR+1}:F{TOT-1})")
c.font, c.alignment, c.border = font(10, True), CTR, BOX
c = sl.cell(row=TOT, column=8, value=f'=COUNTIF(I{SHDR+1}:I{TOT-1},"Done")&" of "&COUNTA(I{SHDR+1}:I{TOT-1})&" done"')
c.font, c.alignment, c.border = font(10, True), CTR, BOX

dv2 = DataValidation(type="list", formula1='"To do,Rendering,Keep,Redo,Done"', allow_blank=True)
sl.add_data_validation(dv2)
dv2.add(f"I{SHDR+1}:I{TOT-1}")

sl.freeze_panes = "B4"
sl.sheet_view.showGridLines = False

# =====================================================================
# 5. MODIFIERS
# =====================================================================
md = wb.create_sheet("Modifiers")
title_row(md, "Modifier library — copy a phrase, drop it into a prompt", 7)
md.cell(row=2, column=1, value=(
    "Reference only; nothing here feeds a formula. One phrase per category per prompt — "
    "stacking three camera moves gets you none of them."
)).font = NOTE_F

COLS = [
    ("Camera angle", ["eye level", "low angle, hero shot", "high angle looking down",
                      "worm's eye view", "over the shoulder", "dutch angle", "top-down",
                      "extreme close-up", "medium shot", "wide establishing shot"]),
    ("Camera move", ["static shot", "slow push in", "slow pull back", "orbit to the left",
                     "tracking shot alongside", "crane up", "tilt up to follow",
                     "handheld follow", "whip pan", "camera tracks backwards"]),
    ("Lighting", ["soft even three-point lighting", "golden hour backlight", "dramatic rim light",
                  "overcast diffused daylight", "neon city glow", "single warm key light",
                  "moonlight, cool blue", "bright flat daylight", "underwater caustics",
                  "firelight from below"]),
    ("Lens / film", ["shallow depth of field", "wide angle lens", "85mm portrait lens",
                     "anamorphic flare", "macro detail", "deep focus", "slight motion blur",
                     "crisp digital sharpness", "soft focus glow", "fisheye"]),
    ("Style", ["bright 3D animated film style, Pixar-like", "hand-drawn 2D cartoon",
               "comic book ink and flat colour", "watercolour storybook",
               "claymation stop-motion", "retro Saturday-morning cartoon",
               "cel-shaded anime", "papercraft cut-out", "chunky toy plastic",
               "crayon drawing brought to life"]),
    ("Setting", ["plain light grey studio backdrop", "city rooftop at sunset",
                 "suburban back garden", "school playground", "coral reef underwater",
                 "cloud tops in daylight", "living room, toys on the floor",
                 "seaside harbour", "night-time high street", "secret base, glowing screens"]),
    ("Motion verb", ["sprints", "leaps", "hovers", "spins", "glides", "skids to a stop",
                     "somersaults", "dives", "soars upward", "lands in a crouch"]),
]

for ci, (head, items) in enumerate(COLS, start=1):
    L = get_column_letter(ci)
    md.column_dimensions[L].width = 30
    c = md.cell(row=3, column=ci, value=head)
    c.font, c.fill, c.alignment, c.border = BAND_F, BAND_FILL, CTR, BOX
    for ri, item in enumerate(items, start=4):
        c = md.cell(row=ri, column=ci, value=item)
        c.font, c.fill, c.alignment, c.border = BODY_F, GREY_FILL, MID, BOX
        md.row_dimensions[ri].height = 18

md.freeze_panes = "A4"
md.sheet_view.showGridLines = False

# =====================================================================
# 6. NEGATIVE PROMPTS
# =====================================================================
ng = wb.create_sheet("Negative Prompts")
widths(ng, {"A": 26, "B": 96, "C": 44})
title_row(ng, "Negative prompts — what to keep out", 3)

ng.cell(row=3, column=1, value="Default (all categories)").font = font(10, True)
ng.cell(row=3, column=1).alignment = MID
ng.cell(row=3, column=3, value="Copy B4 into Kling's negative field. It rebuilds from the rows below.")
ng.cell(row=3, column=3).font = NOTE_F
ng.cell(row=3, column=3).alignment = TOP

NEG = [
    ("Anatomy", "extra fingers, extra limbs, deformed hands, fused fingers, missing limbs, "
                "distorted face, asymmetrical eyes"),
    ("Age drift", "adult body, teenager, elongated adult proportions, mature facial features"),
    ("Continuity", "changed costume, different emblem, colour shift, inconsistent hair, "
                   "wrong eye colour, added accessories"),
    ("Frame", "cropped head, cropped feet, cut off limbs, subject out of frame, "
              "multiple characters, duplicate character"),
    ("Rendering", "blurry, low resolution, jpeg artifacts, oversaturated, heavy grain, "
                  "watermark, text, logo, signature, caption"),
    ("Tone", "scary, gore, blood, weapons, violence, dark horror lighting, menacing expression"),
    ("Motion", "flickering, warping, morphing face, jittery limbs, sudden cuts, "
               "camera shake, stuttering"),
]

r = 6
for col, val in enumerate(["Category", "Terms", "Why"], start=1):
    c = ng.cell(row=r, column=col, value=val)
    c.font, c.fill, c.alignment, c.border = BAND_F, BAND_FILL, CTR, BOX
r += 1
FIRST_NEG = r

WHY = {
    "Anatomy":    "The commonest visible failure, and the one a child notices first.",
    "Age drift":  "Kling's superhero prior is adult. Push back on it in both directions.",
    "Continuity": "The whole reason this workbook exists.",
    "Frame":      "Full-body reference images crop themselves given half a chance.",
    "Rendering":  "Text and watermarks appear from nowhere on comic-styled prompts.",
    "Tone":       "The audience is five. Non-negotiable.",
    "Motion":     "Video only. Harmless on a still, so it stays in the default.",
}
for cat, terms in NEG:
    c = ng.cell(row=r, column=1, value=cat)
    c.font, c.alignment, c.border = font(10, True), TOP, BOX
    c = ng.cell(row=r, column=2, value=terms)
    c.font, c.fill, c.alignment, c.border = INPUT_F, IN_FILL, TOP, BOX
    c = ng.cell(row=r, column=3, value=WHY[cat])
    c.font, c.alignment, c.border = NOTE_F, TOP, BOX
    ng.row_dimensions[r].height = 32
    r += 1
LAST_NEG = r - 1

c = ng.cell(row=4, column=1)
c.border = BOX
c = ng.cell(row=4, column=2, value=f'=_xlfn.TEXTJOIN(", ",TRUE,$B${FIRST_NEG}:$B${LAST_NEG})')
c.font, c.alignment, c.border = FORMULA_F, TOP, BOX
ng.row_dimensions[4].height = 78
c = ng.cell(row=4, column=3, value="Delete a row's terms above and they drop out of here too.")
c.font, c.alignment, c.border = NOTE_F, TOP, BOX

r += 1
ng.cell(row=r, column=1, value="Note").font = font(10, True)
ng.cell(row=r, column=2, value=(
    "Kling weights a long negative prompt weakly — every term dilutes the rest. If one problem "
    "keeps recurring, cut the default down to that problem alone rather than adding to it."
)).font = BODY_F
ng.cell(row=r, column=2).alignment = TOP
ng.row_dimensions[r].height = 32

ng.sheet_view.showGridLines = False

wb.save(OUT)
print("wrote", OUT)
